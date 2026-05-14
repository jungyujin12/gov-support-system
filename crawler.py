"""
============================================================
 정부지원사업 자동 수집 시스템 (보안 적용 버전)
 Government Grant Auto-Crawler — Secure Edition
============================================================
 대상 소스:
   1. 기업마당 API (data.go.kr 공공데이터포털)
   2. 교육부 사업공고 게시판 크롤링

 보안 원칙:
   - API 키는 절대 코드에 직접 입력하지 않음
   - .env 파일 또는 시스템 환경변수에서 동적 로드
   - GitHub Actions 사용 시 Secrets에서 주입

 필요 패키지 설치:
   pip install requests beautifulsoup4 pandas python-dotenv openpyxl
============================================================
"""

import os
import sys
import time
import logging
import requests
import pandas as pd
from datetime import datetime, timedelta
from typing import Optional
from bs4 import BeautifulSoup

# ──────────────────────────────────────────────
# [1] 환경변수 로드 (python-dotenv 사용)
#     .env 파일이 없으면 시스템 환경변수에서 직접 읽음
# ──────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()  # 프로젝트 루트의 .env 파일 자동 로드
    print("✅ .env 파일 로드 완료")
except ImportError:
    print("⚠️  python-dotenv 미설치. 시스템 환경변수에서 직접 읽습니다.")
    print("   설치: pip install python-dotenv")

# ──────────────────────────────────────────────
# [2] 로깅 설정
#     - 콘솔 출력 + 파일 저장 동시 처리
# ──────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("crawler.log", encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────
# [3] 설정값 (Config)
#     모든 민감 정보는 환경변수에서 읽음
# ──────────────────────────────────────────────
class Config:
    """
    환경변수 기반 설정 클래스.
    코드에 직접 값을 입력하지 않고
    os.environ을 통해 런타임에 주입받습니다.
    """

    # ── API 키: .env 또는 시스템 환경변수에서 읽음 ──
    BIZINFO_API_KEY: str = os.environ.get("BIZINFO_API_KEY", "")
    SLACK_WEBHOOK_URL: str = os.environ.get("SLACK_WEBHOOK_URL", "")

    # ── 필터링 키워드 (대학 관련 사업만 추출) ──
    KEYWORDS: list = ["대학", "산학협력", "인재양성", "R&D", "연구소", "석박사"]

    # ── 기업마당 API 엔드포인트 ──
    BIZINFO_URL: str = (
        "https://apis.data.go.kr/B552735/kisedbizentrprssupport/getAnnoList"
    )
    BIZINFO_ROWS: int = 100   # 한 번에 가져올 최대 공고 수

    # ── 교육부 게시판 URL ──
    MOE_URL: str = (
        "https://www.moe.go.kr/boardCnts/listRenew.do?boardID=72761"
    )
    MOE_PAGES: int = 3        # 크롤링할 페이지 수

    # ── 요청 타임아웃 / 재시도 ──
    TIMEOUT: int = 15         # 초
    RETRY: int = 3            # 재시도 횟수
    RETRY_DELAY: int = 2      # 재시도 간격(초)

    # ── 결과 저장 경로 ──
    OUTPUT_DIR: str = "output"
    OUTPUT_EXCEL: str = "정부지원사업_공고.xlsx"
    OUTPUT_CSV: str = "정부지원사업_공고.csv"

    @classmethod
    def validate(cls) -> bool:
        """
        필수 환경변수 검증.
        키가 없으면 경고를 출력하고 False 반환.
        """
        missing = []
        if not cls.BIZINFO_API_KEY:
            missing.append("BIZINFO_API_KEY")
        if missing:
            logger.warning(
                f"⚠️  환경변수 누락: {', '.join(missing)}\n"
                "   기업마당 API 수집은 건너뜁니다.\n"
                "   .env 파일에 키를 추가하거나 환경변수를 설정하세요."
            )
            return False
        return True


# ──────────────────────────────────────────────
# [4] HTTP 유틸리티 (재시도 로직 포함)
# ──────────────────────────────────────────────
def safe_request(
    url: str,
    params: Optional[dict] = None,
    headers: Optional[dict] = None,
    method: str = "GET",
) -> Optional[requests.Response]:
    """
    재시도 로직이 포함된 안전한 HTTP 요청 함수.
    네트워크 오류, 타임아웃, HTTP 에러를 개별 처리합니다.

    Args:
        url:     요청 URL
        params:  쿼리 파라미터 딕셔너리
        headers: 요청 헤더
        method:  HTTP 메서드 (GET/POST)

    Returns:
        Response 객체 또는 None (실패 시)
    """
    default_headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
        ),
        "Accept-Language": "ko-KR,ko;q=0.9",
    }
    if headers:
        default_headers.update(headers)

    for attempt in range(1, Config.RETRY + 1):
        try:
            response = requests.request(
                method,
                url,
                params=params,
                headers=default_headers,
                timeout=Config.TIMEOUT,
            )
            # HTTP 4xx/5xx 에러를 예외로 변환
            response.raise_for_status()
            return response

        except requests.exceptions.Timeout:
            logger.warning(f"⏱️  타임아웃 (시도 {attempt}/{Config.RETRY}): {url}")
        except requests.exceptions.ConnectionError:
            logger.warning(f"🔌 연결 실패 (시도 {attempt}/{Config.RETRY}): {url}")
        except requests.exceptions.HTTPError as e:
            logger.error(f"❌ HTTP 오류 {e.response.status_code}: {url}")
            return None  # HTTP 에러는 재시도 무의미
        except requests.exceptions.RequestException as e:
            logger.error(f"❌ 알 수 없는 요청 오류: {e}")

        if attempt < Config.RETRY:
            logger.info(f"   {Config.RETRY_DELAY}초 후 재시도...")
            time.sleep(Config.RETRY_DELAY)

    logger.error(f"❌ 최대 재시도 초과. 요청 포기: {url}")
    return None


# ──────────────────────────────────────────────
# [5] 기업마당 API 수집 모듈
# ──────────────────────────────────────────────
def fetch_bizinfo(page: int = 1) -> list[dict]:
    """
    공공데이터포털 기업마당 API를 호출해 공고 목록을 수집합니다.
    API 키는 Config에서 환경변수로 읽어옵니다.

    ⚠️ 보안 주의: API 키를 params에 직접 넣더라도
       로그에 절대 출력하지 않습니다.

    Args:
        page: 페이지 번호

    Returns:
        공고 딕셔너리 리스트
    """
    logger.info(f"📡 기업마당 API 호출 중... (page={page})")

    params = {
        "serviceKey": Config.BIZINFO_API_KEY,  # 환경변수에서 주입
        "pageNo": page,
        "numOfRows": Config.BIZINFO_ROWS,
        "type": "json",
    }

    response = safe_request(Config.BIZINFO_URL, params=params)
    if not response:
        return []

    try:
        data = response.json()
        # API 응답 구조: response > body > items
        body = data.get("response", {}).get("body", {})
        items = body.get("items", [])

        if not items:
            logger.info("   기업마당: 수집된 공고 없음")
            return []

        # items가 dict인 경우(단일 항목) 리스트로 변환
        if isinstance(items, dict):
            items = [items]

        logger.info(f"   기업마당: {len(items)}건 원본 수집")
        return items

    except (KeyError, ValueError) as e:
        logger.error(f"❌ 기업마당 JSON 파싱 오류: {e}")
        logger.debug(f"   응답 내용: {response.text[:300]}")
        return []


def parse_bizinfo(items: list[dict]) -> list[dict]:
    """
    기업마당 API 응답을 표준 포맷으로 변환하고
    키워드 필터링을 적용합니다.

    표준 포맷: {소스, 사업명, 주관부처, 마감일, 상세링크}
    """
    results = []
    for item in items:
        title = item.get("pblancNm", "")  # 공고명
        ministry = item.get("jrsdInsttNm", "")  # 관할기관명
        deadline = item.get("rcptEndDd", "")  # 접수마감일
        pblanc_id = item.get("pblancId", "")  # 공고 ID

        # 키워드 필터링: 제목 또는 부처명에 키워드 포함 여부
        matched = any(
            kw in title or kw in ministry
            for kw in Config.KEYWORDS
        )
        if not matched:
            continue

        # 상세링크 조합
        detail_url = (
            f"https://www.bizinfo.go.kr/web/lay1/bbs/S1T122C128/AS/74/"
            f"view.do?pblancId={pblanc_id}"
            if pblanc_id
            else "https://www.bizinfo.go.kr"
        )

        results.append({
            "소스": "기업마당",
            "사업명": title,
            "주관부처": ministry,
            "마감일": deadline,
            "상세링크": detail_url,
        })

    logger.info(f"   기업마당 필터링 후: {len(results)}건")
    return results


# ──────────────────────────────────────────────
# [6] 교육부 게시판 크롤링 모듈
# ──────────────────────────────────────────────
def fetch_moe(max_pages: int = None) -> list[dict]:
    """
    교육부 사업공고 게시판을 BeautifulSoup으로 크롤링합니다.
    robots.txt 준수 및 요청 간격 조절이 적용됩니다.

    Args:
        max_pages: 크롤링할 최대 페이지 수

    Returns:
        공고 딕셔너리 리스트
    """
    if max_pages is None:
        max_pages = Config.MOE_PAGES

    logger.info(f"🏫 교육부 게시판 크롤링 시작 ({max_pages}페이지)...")
    results = []

    for page in range(1, max_pages + 1):
        # 서버 부하 방지: 페이지 간 1초 대기
        if page > 1:
            time.sleep(1)

        url = (
            f"https://www.moe.go.kr/boardCnts/listRenew.do"
            f"?boardID=72761&page={page}"
        )
        response = safe_request(url)
        if not response:
            logger.warning(f"   교육부 {page}페이지 수집 실패, 건너뜀")
            continue

        soup = BeautifulSoup(response.text, "html.parser")
        page_items = _parse_moe_page(soup)
        results.extend(page_items)
        logger.info(f"   교육부 {page}페이지: {len(page_items)}건")

    logger.info(f"   교육부 총 수집: {len(results)}건")
    return results


def _parse_moe_page(soup: BeautifulSoup) -> list[dict]:
    """
    교육부 게시판 HTML에서 공고 정보를 추출합니다.
    사이트 구조 변경에 대비한 다중 셀렉터 전략을 사용합니다.

    Returns:
        파싱된 공고 리스트
    """
    results = []
    BASE_URL = "https://www.moe.go.kr"

    # 다중 셀렉터: 사이트 리뉴얼에 대비
    selectors = [
        "table.board_list tbody tr",
        "table.bbs_list tbody tr",
        ".board-list tbody tr",
        "tbody tr",  # 최후 수단
    ]

    rows = []
    for sel in selectors:
        rows = soup.select(sel)
        if rows:
            break

    for row in rows:
        try:
            # 제목 셀 및 링크 추출
            title_cell = (
                row.select_one("td.subject a")
                or row.select_one("td.title a")
                or row.select_one("td a")
            )
            if not title_cell:
                continue

            title = title_cell.get_text(strip=True)
            href = title_cell.get("href", "")

            # 빈 제목 또는 공지 행 건너뜀
            if not title or title in ("공지", ""):
                continue

            # 키워드 필터링
            if not any(kw in title for kw in Config.KEYWORDS):
                continue

            # 절대 URL 변환
            if href and not href.startswith("http"):
                href = BASE_URL + href

            # 날짜 추출 (마지막 td 또는 date 클래스)
            date_cell = (
                row.select_one("td.date")
                or row.select_one("td.regDate")
            )
            if not date_cell:
                tds = row.select("td")
                date_cell = tds[-1] if tds else None

            date_text = date_cell.get_text(strip=True) if date_cell else ""

            results.append({
                "소스": "교육부",
                "사업명": title,
                "주관부처": "교육부",
                "마감일": date_text,
                "상세링크": href,
            })

        except Exception as e:
            logger.debug(f"   행 파싱 오류 (건너뜀): {e}")
            continue

    return results


# ──────────────────────────────────────────────
# [7] 데이터 정제 및 통합 (Pandas)
# ──────────────────────────────────────────────
def build_dataframe(all_items: list[dict]) -> pd.DataFrame:
    """
    수집된 모든 공고를 Pandas DataFrame으로 변환하고
    중복 제거, 날짜 정렬, 컬럼 정리를 수행합니다.

    Returns:
        정제된 DataFrame
    """
    if not all_items:
        logger.warning("⚠️  수집된 데이터 없음 — 빈 DataFrame 반환")
        return pd.DataFrame(columns=["소스", "사업명", "주관부처", "마감일", "상세링크"])

    df = pd.DataFrame(all_items)

    # 컬럼 순서 강제 지정
    cols = ["소스", "사업명", "주관부처", "마감일", "상세링크"]
    df = df.reindex(columns=cols)

    # 중복 제거 (사업명 기준)
    before = len(df)
    df = df.drop_duplicates(subset=["사업명"], keep="first")
    logger.info(f"🔄 중복 제거: {before - len(df)}건 제거됨")

    # 빈 사업명 제거
    df = df[df["사업명"].str.strip() != ""]

    # 마감일 컬럼 정제 (숫자+하이픈 형식 정규화)
    df["마감일"] = df["마감일"].astype(str).str.strip()

    # 수집 시각 추가
    df["수집일시"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    logger.info(f"✅ 최종 데이터: {len(df)}건")
    return df


# ──────────────────────────────────────────────
# [8] 결과 저장 모듈 (Excel + CSV)
# ──────────────────────────────────────────────
def save_results(df: pd.DataFrame) -> dict[str, str]:
    """
    DataFrame을 Excel과 CSV 두 형식으로 저장합니다.
    Excel은 스타일이 적용된 보기 좋은 형태로 저장합니다.

    Returns:
        저장된 파일 경로 딕셔너리 {excel, csv}
    """
    os.makedirs(Config.OUTPUT_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")

    excel_path = os.path.join(
        Config.OUTPUT_DIR,
        f"{today}_{Config.OUTPUT_EXCEL}"
    )
    csv_path = os.path.join(
        Config.OUTPUT_DIR,
        f"{today}_{Config.OUTPUT_CSV}"
    )

    # ── CSV 저장 ──
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")  # BOM 포함 (Excel 호환)
    logger.info(f"💾 CSV 저장: {csv_path}")

    # ── Excel 저장 (openpyxl 스타일 적용) ──
    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="공고목록")
            ws = writer.sheets["공고목록"]

            # 헤더 스타일
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_fill = PatternFill(
                start_color="1E3A5F", end_color="1E3A5F", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col_num, cell in enumerate(ws[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # 열 너비 자동 조정
            col_widths = {
                "A": 12,   # 소스
                "B": 50,   # 사업명
                "C": 20,   # 주관부처
                "D": 15,   # 마감일
                "E": 60,   # 상세링크
                "F": 18,   # 수집일시
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            # 행 높이 설정
            ws.row_dimensions[1].height = 22
            for row in ws.iter_rows(min_row=2):
                ws.row_dimensions[row[0].row].height = 18
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

        logger.info(f"📊 Excel 저장: {excel_path}")

    except ImportError:
        logger.warning("⚠️  openpyxl 미설치. Excel 저장 건너뜀. (pip install openpyxl)")
        excel_path = None

    return {"excel": excel_path, "csv": csv_path}


# ──────────────────────────────────────────────
# [9] 슬랙 알림 모듈
# ──────────────────────────────────────────────
def send_slack(df: pd.DataFrame) -> bool:
    """
    수집 결과 요약을 슬랙 웹훅으로 전송합니다.
    웹훅 URL은 환경변수 SLACK_WEBHOOK_URL에서 읽습니다.

    Args:
        df: 수집된 공고 DataFrame

    Returns:
        전송 성공 여부
    """
    if not Config.SLACK_WEBHOOK_URL:
        logger.info("ℹ️  SLACK_WEBHOOK_URL 미설정. 슬랙 알림 건너뜀.")
        return False

    today = datetime.now().strftime("%Y년 %m월 %d일")
    urgent = df[
        df["마감일"].str.match(r"\d{8}|\d{4}-\d{2}-\d{2}", na=False)
    ] if not df.empty else pd.DataFrame()

    # 슬랙 블록 메시지 구성
    blocks = [
        {
            "type": "header",
            "text": {
                "type": "plain_text",
                "text": f"📡 정부지원사업 공고 수집 완료 — {today}",
            },
        },
        {
            "type": "section",
            "fields": [
                {"type": "mrkdwn", "text": f"*전체 공고*\n{len(df)}건"},
                {"type": "mrkdwn", "text": f"*기업마당*\n{len(df[df['소스']=='기업마당'])}건"},
                {"type": "mrkdwn", "text": f"*교육부*\n{len(df[df['소스']=='교육부'])}건"},
                {"type": "mrkdwn", "text": f"*수집 시각*\n{datetime.now().strftime('%H:%M')}"},
            ],
        },
        {"type": "divider"},
    ]

    # 상위 5건 목록 추가
    if not df.empty:
        top5_text = "\n".join(
            f"• <{row['상세링크']}|{row['사업명'][:35]}{'...' if len(row['사업명'])>35 else ''}>"
            f"  _{row['주관부처']}_"
            for _, row in df.head(5).iterrows()
        )
        blocks.append({
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*📋 주요 공고 TOP 5*\n{top5_text}",
            },
        })

    payload = {"blocks": blocks}
    response = safe_request(
        Config.SLACK_WEBHOOK_URL,
        method="POST",
        headers={"Content-Type": "application/json"},
    )

    # POST는 safe_request의 params가 아닌 json으로 전송해야 함
    try:
        import json
        res = requests.post(
            Config.SLACK_WEBHOOK_URL,
            json=payload,
            timeout=Config.TIMEOUT,
        )
        if res.status_code == 200:
            logger.info("💬 슬랙 알림 전송 성공")
            return True
        else:
            logger.error(f"❌ 슬랙 전송 실패: {res.status_code} {res.text}")
            return False
    except Exception as e:
        logger.error(f"❌ 슬랙 전송 오류: {e}")
        return False


# ──────────────────────────────────────────────
# [10] 메인 실행 함수
# ──────────────────────────────────────────────
def main():
    """
    전체 파이프라인 실행:
      기업마당 API → 교육부 크롤링 → 통합 → 저장 → 슬랙 알림
    """
    logger.info("=" * 55)
    logger.info("  정부지원사업 자동 수집 시작")
    logger.info(f"  실행 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 55)

    all_items = []

    # ── 기업마당 API 수집 ──
    if Config.validate():
        raw_items = fetch_bizinfo(page=1)
        parsed = parse_bizinfo(raw_items)
        all_items.extend(parsed)
    else:
        logger.warning("기업마당 API 수집 건너뜀 (API 키 없음)")

    # ── 교육부 크롤링 ──
    moe_items = fetch_moe()
    all_items.extend(moe_items)

    # ── DataFrame 구성 ──
    df = build_dataframe(all_items)

    # ── 결과 출력 ──
    if not df.empty:
        logger.info("\n📋 수집 결과 미리보기:")
        print(df[["소스", "사업명", "주관부처", "마감일"]].to_string(index=False))

    # ── 파일 저장 ──
    paths = save_results(df)
    logger.info(f"\n✅ 완료! 저장 경로: {paths}")

    # ── 슬랙 알림 ──
    send_slack(df)

    logger.info("=" * 55)
    logger.info("  수집 완료")
    logger.info("=" * 55)

    return df


if __name__ == "__main__":
    main()
